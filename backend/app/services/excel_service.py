from io import BytesIO

import openpyxl

from sqlalchemy.orm import Session

from app.models.contact_file import ContactFile
from app.models.contact import Contact


def import_contacts_from_excel(
    db: Session,
    file,
    company_id: str,
    imported_by: str,
    filename: str
):
    """
    Expected sheet structure (row 1 = header, skipped):
        column A: name
        column B: email
        column C: department

    Contacts are unique per company by (name, email). Duplicates
    (already in the DB, or repeated within the same file) are skipped.
    """

    # Read the whole upload into memory and wrap it in BytesIO so
    # openpyxl gets a proper seekable binary stream regardless of what
    # kind of file-like object FastAPI handed us.
    file.seek(0)
    contents = file.read()

    if not contents:
        raise ValueError(
            "The uploaded file is empty (0 bytes). Please re-select the "
            "file and try again."
        )

    try:
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise ValueError(
            f"Could not read this file as an Excel (.xlsx) file: {exc}"
        ) from exc

    if not workbook.sheetnames:
        raise ValueError("Excel file has no sheets")

    sheet = workbook.active

    # Create the file record first so contacts can reference it
    contact_file = ContactFile(
        company_id=company_id,
        filename=filename
    )

    db.add(contact_file)
    db.flush()

    # Existing (name, email) pairs already stored for this company
    existing_pairs = {
        (name.strip().lower(), email.strip().lower())
        for (name, email) in (
            db.query(Contact.name, Contact.email)
            .filter(Contact.company_id == company_id)
            .all()
        )
        if name and email
    }

    imported = 0
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row or len(row) < 2:
            continue

        name = row[0]
        email = row[1]
        department = row[2] if len(row) > 2 else None

        if not name or not email:
            continue

        name = str(name).strip()
        email = str(email).strip()
        key = (name.lower(), email.lower())

        if key in existing_pairs:
            skipped += 1
            continue

        existing_pairs.add(key)

        contact = Contact(
            company_id=company_id,
            imported_by=imported_by,
            file_id=contact_file.id_file,
            name=name,
            email=email,
            department=str(department).strip() if department else None,
        )

        db.add(contact)
        imported += 1

    db.commit()
    db.refresh(contact_file)

    return contact_file, imported, skipped